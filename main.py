from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
import random
import math
rand = 0
G = 6.67 * 10 ** (-11)
RadiusSun = 6.9551 * 10 ** 8
RadiusEarth = 6371 * 1000
MassSun = 1.9891 * 10 ** 30
MassEarth = 5.97 * 10 ** 24
LuminositySun = 3.939 * 10 ** 26
PSB = 5.67 * 10 ** (-8)
AE = 149600000000
myFill = PatternFill(start_color='d0cece', end_color='d0cece', fill_type='solid')
StringNames = ['Название', 'Тип', 'Радиуc, м', 'Радиуc, Rз', 'Масса, кг', 'Масса, Мз', 'Плотность, кг/м^3',
               'Ускорение свободного падения, м/с^2', 'Большая полуось орбиты, м', 'Большая полуось орбиты, а.е.',
               'Период вращения, дни', 'Период обращения, дни', 'Уровень взаимодействия', '1 космическая скорость, м/с',
               '2 космическая скорость, м/с', 'Скорость движения по орбите, м/с', 'Угол наклона оси, градусы']
TypesStar = ['Горячая звезда', 'Горячий сверхгигант', 'Звезда', 'Гигант', 'Сверхгигант',
             'Коричневый карлик', 'Белый карлик', 'Чёрный карлик', 'Пульсар', 'Магнитар']
ListKlassPlanetHot = ['Каменная планета', 'Насыщенная металлами планета', 'Металлическая планета',
                  'Безоблачный гигант', 'Щелочной гигант', 'Кремневый гигант']
ListKlassPlanetVen = ['Каменная планета', 'Насыщенная металлами планета', 'Металлическая планета',
                      'Венероподобная планета']
ListKlassPlanetLife = ['Каменная планета', 'Насыщенная металлами планета', 'Металлическая планета',
                      'Водная планета', 'Землеподобная планета', 'Водный гигант']
ListKlassPlanetCold = ['Каменная планета', 'Насыщенная металлами планета', 'Металлическая планета',
                      'Ледяная планета', 'Каменно-ледяная планета', 'Амиачная планета', 'Амиачный гигант',
                       'Ледяной гигант']
TypesPlanet = ['Нептуноподобная', 'Юпитероподобная', 'Планета земной группы', 'Суперземля']
ListRings = ['Есть', 'Нет']
TypesMoon = ['Каменный', 'Насыщенный металлами', 'Металлический', 'Водный', 'Землеподобный',
             'Ледяной', 'Каменно-ледяной', 'Амиачный']


def CountStar():
    global rand
    if rand == -1:
        inp = -1
    else:
        inp = int(input("Сколько звезд в системе? "))
        if inp == -2:
            rand = -1
            inp = -1
    if inp == -1:
        return random.choice(list([8]*2+[7]*4+[6]*5+[5]*10+[4]*20+[3]*60+[2]*100+[1]*799))
    else:
        return inp

def TypeStar(name_star):
    global TypesStar, rand
    if rand == -1:
        inp = -1
    else:
        print("Выберете тип " + name_star)
        for _ in range(1, 11):
            print(str(_) + '.' + TypesStar[_ - 1])
        inp = int(input())
        if inp == -2:
            rand = -1
            inp = -1
    if inp == -1:
        return random.choice(list([1]*5+[2]*1+[3]*74+[4]*4+[5]*2+[6]*5+[7]*5+[8]*1+[9]*2+[10]))
    else:
        return inp

def CountPlanet(name_star):
    global rand
    if rand == -1:
        inp = -1
    else:
        inp = int(input("Сколько планет у звезды "+name_star+"? "))
        if inp == -2:
            rand = -1
            inp = -1
    if inp == -1:
        return random.choice(list([0,1,2,3,3,4,4,5,5,5,6,6,6,7,7,7,8,8,9,9,10,11,12,13,14,15,16,17,18,19,20]))
    else:
        return inp

def KlassStar(name_star, type_star):
    global rand
    if rand == -1:
        inp = '-1'
    else:
        inp = 0
    if type_star == 1:
        if inp != '-1':
            print("Выберете класс "+name_star)
            print('O\nB\nA\nF')
            inp = input()
        if inp == '-2':
            rand = -1
            inp = '-1'
        if inp == '-1':
            return random.choice(list(['F']*71+['A']*23+['B']*4+['O']*2))
        else:
            return inp
    elif type_star == 2:
        if inp != '-1':
            print("Выберете класс " + name_star)
            print('Oc\nBc\nAc\nFc')
            inp = input()
        if inp == '-2':
            rand = -1
            inp = '-1'
        if inp == '-1':
            return random.choice(list(['Fc']*71+['Ac']*23+['Bc']*4+['Oc']*2))
        else:
            return inp
    elif type_star == 3:
        if inp != '-1':
            print("Выберете класс " + name_star)
            print('G\nK\nM')
            inp = input()
        if inp == '-2':
            rand = -1
            inp = '-1'
        if inp == '-1':
            return random.choice(list(['G']*4+['K']*9+['M']*87))
        else:
            return inp
    elif type_star == 4:
        if inp != '-1':
            print("Выберете класс " + name_star)
            print('Gg\nKg\nMg')
            inp = input()
        if inp == '-2':
            rand = -1
            inp = '-1'
        if inp == '-1':
            return random.choice(list(['Gg'] * 4 + ['Kg'] * 9 + ['Mg'] * 87))
        else:
            return inp
    elif type_star == 5:
        if inp != '-1':
            print("Выберете класс " + name_star)
            print('Gc\nKc\nMc')
            inp = input()
        if inp == '-2':
            rand = -1
            inp = '-1'
        if inp == '-1':
            return random.choice(list(['Gc'] * 4 + ['Kc'] * 9 + ['Mc'] * 87))
        else:
            return inp
    elif type_star == 6:
        if inp != '-1':
            print("Выберете класс " + name_star)
            print('L\nT\nY')
            inp = input()
        if inp == '-2':
            rand = -1
            inp = '-1'
        if inp == '-1':
            return random.choice(list(['L']*6+['T']*6+['Y']))
        else:
            return inp
    elif type_star == 7:
        return 'D'
    elif type_star == 8:
        return 'Dd'
    else:
        return 'N'

def ColorStar(klass_star):
    if klass_star == 'O' or klass_star == 'Oc':
        return 'Голубой'
    elif klass_star == 'B' or klass_star == 'Bc':
        return 'Бело-голубой'
    elif klass_star == 'A' or klass_star == 'Ac' or klass_star == 'D':
        return 'Белый'
    elif klass_star == 'F' or klass_star == 'Fc':
        return 'Жёлто-белый'
    elif klass_star == 'G' or klass_star == 'Gc' or klass_star == 'Gg':
        return 'Жёлтый'
    elif klass_star == 'K' or klass_star == 'Kc' or klass_star == 'Kg':
        return 'Оранжевый'
    elif klass_star == 'M' or klass_star == 'Mc' or klass_star == 'Mg':
        return 'Красный'
    elif klass_star == 'L' or klass_star == 'T' or klass_star == 'Y':
        return 'Коричневый'
    elif klass_star == 'Dd':
        return 'Чёрный'
    elif klass_star == 'N':
        return 'Фиолетово-белый'

def TempStar(klass_star):
    if klass_star == 'O' or klass_star == 'Oc':
        return random.randint(300, 600)*100
    elif klass_star == 'B' or klass_star == 'Bc':
        return random.randint(100, 299)*100
    elif klass_star == 'A' or klass_star == 'Ac':
        return random.randint(750, 999)*10
    elif klass_star == 'F' or klass_star == 'Fc':
        return random.randint(600, 749)*10
    elif klass_star == 'G' or klass_star == 'Gc' or klass_star == 'Gg':
        return random.randint(500, 599)*10
    elif klass_star == 'K' or klass_star == 'Kc' or klass_star == 'Kg':
        return random.randint(350, 499)*10
    elif klass_star == 'M' or klass_star == 'Mc' or klass_star == 'Mg':
        return random.randint(200, 349)*10
    elif klass_star == 'L':
        return random.randint(130, 299)*10
    elif klass_star == 'T':
        return random.randint(70, 129)*10
    elif klass_star == 'Y':
        return random.randint(20, 69)*10
    elif klass_star == 'Dd':
        return random.randint(1, 100)
    elif klass_star == 'D':
        return random.randint(60, 200)*100
    elif klass_star == 'N':
        return random.randint(100, 1000)*1000

def RadiusStar(klass_star):
    if klass_star == 'O':
        inp = random.randint(105, 195)/10
    elif klass_star == 'B':
        inp = random.randint(49, 91)/10
    elif klass_star == 'A':
        inp = random.randint(15, 27)/10
    elif klass_star == 'F':
        inp = random.randint(9, 17)/10
    elif klass_star == 'G':
        inp = random.randint(77, 140)/100
    elif klass_star == 'K':
        inp = random.randint(60, 110)/100
    elif klass_star == 'M':
        inp = random.randint(9, 71)/100
    elif klass_star == 'L' or klass_star == 'T' or klass_star == 'Y':
        inp = random.randint(12, 70)/1000
    elif klass_star == 'Dd' or klass_star == 'D':
        return '-', random.randint(609, 819)*1000
    elif klass_star == 'N':
        return '-', random.randint(100, 150)*100
    elif len(klass_star) == 2 and klass_star[1] == 'g':
        inp = random.randint(10, 100)
    elif len(klass_star) == 2 and klass_star[1] == 'c':
        inp = random.randint(10, 180)*10
    return inp, inp * RadiusSun

def MassStar(klass_star):
    if klass_star == 'O':
        inp = random.randint(15, 105)
    elif klass_star == 'B':
        inp = random.randint(45, 315)/10
    elif klass_star == 'A':
        inp = random.randint(78, 543)/100
    elif klass_star == 'F':
        inp = random.randint(43, 298)/100
    elif klass_star == 'G':
        inp = random.randint(28, 193)/100
    elif klass_star == 'K':
        inp = random.randint(20, 140)/100
    elif klass_star == 'M':
        inp = random.randint(8, 60)/100
    elif klass_star == 'L' or klass_star == 'T' or klass_star == 'Y':
        inp = random.randint(20, 77)/1000
    elif klass_star == 'Dd' or klass_star == 'D':
        inp = random.randint(60, 144)/100
    elif klass_star == 'N':
        inp = random.randint(1, 216)/100
    elif len(klass_star) == 2 and klass_star[1] == 'g':
        inp = random.randint(2, 300)/10
    elif len(klass_star) == 2 and klass_star[1] == 'c':
        inp = random.randint(10, 300)
    return inp, inp * MassSun

def PlanetOrbit(o0):
    if o0 == 0:
        o0 = random.choice(list([random.randint(306, 380)] + [random.randint(232, 306)] * 2 +
                                [random.randint(158, 232)] * 4 + [random.randint(84, 158)] * 8 +
                                [random.randint(10, 84)] * 16))/1000
    else:
        o0 = o0 * random.choice(list([random.randint(317, 358)] + [random.randint(277, 317)] * 2 +
                                [random.randint(236, 277)] * 4 + [random.randint(195, 236)] * 8 +
                                [random.randint(155, 195)] * 16 + [random.randint(114, 155)] * 16))/100
    return o0, o0 * AE

def KlassPlanet(ven1, ven2, life1, life2, orbit, name_planet):
    global rand
    if rand == -1:
        inp = -1
    else:
        inp = 0
    if orbit < ven1:
        if inp != -1:
            print("Выберете класс планеты " + name_planet)
            for _ in range(6):
                print(str(_+1) + '.' + ListKlassPlanetHot[_])
            inp = int(input())
        if inp == -2:
            rand = -1
            inp = -1
        if inp == -1:
            inp = random.choice(list([0] * 4 + [1] * 3 + [2] * 1 + [3] * 6 + [4] * 5 + [5] * 2))
        else:
            inp -= 1
        return ListKlassPlanetHot[inp]
    elif orbit >= ven1 and orbit < ven2:
        if inp != -1:
            print("Выберете класс планеты " + name_planet)
            for _ in range(4):
                print(str(_ + 1) + '.' + ListKlassPlanetVen[_])
            inp = int(input())
        if inp == -2:
            rand = -1
            inp = -1
        if inp == -1:
            inp = random.choice(list([0] * 4 + [1] * 3 + [2] * 1 + [3] * 6))
        else:
            inp -= 1
        return ListKlassPlanetVen[inp]
    elif orbit >= life1 and orbit <= life2:
        if inp != -1:
            print("Выберете класс планеты " + name_planet)
            for _ in range(6):
                print(str(_ + 1) + '.' + ListKlassPlanetLife[_])
            inp = int(input())
        if inp == -2:
            rand = -1
            inp = -1
        if inp == -1:
            inp = random.choice(list([0] * 3 + [1] * 2 + [2] * 1 + [3] * 6 + [4] * 6 + [5] * 5))
        else:
            inp -= 1
        return ListKlassPlanetLife[inp]
    elif orbit > life2:
        if inp != -1:
            print("Выберете класс планеты " + name_planet)
            for _ in range(8):
                print(str(_ + 1) + '.' + ListKlassPlanetCold[_])
            inp = int(input())
        if inp == -2:
            rand = -1
            inp = -1
        if inp == -1:
            inp = random.choice(list([0] * 8 + [1] * 6 + [2] * 3 + [3] * 8 + [4] * 5 + [5] * 2 + [6] * 10 + [7] * 26))
        else:
            inp -= 1
        return ListKlassPlanetCold[inp]

def TypePlanet(klass, name_planet):
    global rand
    if rand == -1:
        inp = -1
    else:
        inp = 0
    if 'гигант' in klass:
        if inp != -1:
            print("Выберете тип планеты " + name_planet)
            print('1.Нептуноподобная\n2.Юпитероподобная')
            inp = int(input())
        if inp == -2:
            rand = -1
            inp = -1
        if inp == -1:
            inp = random.choice(list([0] * 7 + [1] * 4))
        else:
            inp -= 1
        return TypesPlanet[inp]
    else:
        if inp != -1:
            print("Выберете тип планеты " + name_planet)
            print('3.Планета земной группы\n4.Суперземля')
            inp = int(input())
        if inp == -2:
            rand = -1
            inp = -1
        if inp == -1:
            inp = random.choice(list([2] * 1 + [3] * 2))
        else:
            inp -= 1
        return TypesPlanet[inp]

def CountMoon(type, name_planet):
    global rand
    if rand == -1:
        inp = -1
    else:
        inp = 0
    if type[0] == 'П':
        if inp != -1:
            print("Сколько крупных спутников у планеты " + name_planet+'?')
            inp = int(input())
        if inp == -2:
            rand = -1
            inp = -1
        if inp == -1:
            nb = random.choice(list([0] * 5 + [1] * 7 + [2] * 2 + [3]))
        else:
            nb = inp
        ri = ListRings[random.choice(list([0] * 1 + [1] * 19))]
        n = random.randint(nb, max(nb, 11))
        return ri, n, nb
    elif type[0] == 'С':
        if inp != -1:
            print("Сколько крупных спутников у планеты " + name_planet+'?')
            inp = int(input())
        if inp == -2:
            rand = -1
            inp = -1
        if inp == -1:
            nb = random.choice(list([0] * 3 + [1] * 5 + [2] * 3 + [3] * 2 + [4]))
        else:
            nb = inp
        ri = ListRings[random.choice(list([0] * 8 + [1] * 12))]
        n = random.randint(nb, max(nb, 17))
        return ri, n, nb
    elif type[0] == 'Н':
        if inp != -1:
            print("Сколько крупных спутников у планеты " + name_planet+'?')
            inp = int(input())
        if inp == -2:
            rand = -1
            inp = -1
        if inp == -1:
            nb = random.choice(list([0] * 1 + [1] * 1 + [2] * 2 + [3] * 2 + [4] * 2 + [5] * 3 + [6] * 3 +
                                    [7] * 2 + [8] + [9] + [10]))
        else:
            nb = inp
        ri = ListRings[random.choice(list([0] * 14 + [1] * 6))]
        n = random.randint(nb, max(nb, 41))
        return ri, n, nb
    elif type[0] == 'Ю':
        if inp != -1:
            print("Сколько крупных спутников у планеты " + name_planet+'?')
            inp = int(input())
        if inp == -2:
            rand = -1
            inp = -1
        if inp == -1:
            nb = random.choice(list([0] * 1 + [1] * 1 + [2] * 2 + [3] * 2 + [4] * 2 + [5] * 3 + [6] * 3 +
                                    [7] * 2 + [8] + [9] + [10]))
        else:
            nb = inp
        ri = ListRings[random.choice(list([0] * 18 + [1] * 2))]
        n = random.randint(nb, max(nb, 41))
        return ri, n, nb

def MassPlanet(type, star_mass):
    if type[0] == 'П':
        mass = random.randint(2, 190) / 100
        density = random.randint(110, 700) * 10
    elif type[0] == 'С':
        mass = random.randint(19, 340) / 10
        density = random.randint(110, 800) * 10
    elif type[0] == 'Н':
        mass = random.randint(30, 559) / 10
        density = random.randint(50, 180) * 10
    elif type[0] == 'Ю':
        mass = random.randint(56, min(star_mass * 332940 // 3, 5072)) / 10
        density = random.randint(50, 180) * 10
    return density, mass, mass * MassEarth

def TypeMoon(ven1, ven2, life1, life2, orbit):
    if orbit < ven2:
        inp = random.choice(list([0] * 4 + [1] * 2 + [2] * 1))
    elif orbit >= life1 and orbit <= life2:
        inp = random.choice(list([0] * 12 + [1] * 6 + [2] * 3 + [3] + [4]))
    elif orbit > life2:
        inp = random.choice(list([0] * 4 + [1] * 2 + [2] * 1 + [5] * 3 + [6] * 3 + [7]))
    return TypesMoon[inp]




random.seed()
tmp_yn = input("Сгенерировать название системы(y/n)? ")
if tmp_yn == 'y':
    name = ""
    data_syllables = ['ал', 'ле', 'лу', 'ге', 'за', 'се', 'на', 'би', 'со', 'ус', 'юс', 'ес', 'ар', 'ма', 'ин', 'ди',
                      'ре', 'ри', 'а', 'ер', 'ат', 'ен', 'бе', 'ра', 'ла', 'ве', 'ти', 'ед', 'ор', 'ку', 'ан', 'те',
                      'ис', 'он', 'фо', 'ке', 'ек', 'ос']
    tmp_count = random.choice([2, 3, 3, 4, 4, 5])
    for i in range(tmp_count):
        if random.randint(0, 1) == 0:
            name += random.choice(data_syllables)[::-1]
        else:
            name += random.choice(data_syllables)
    name = name.title()
else:
    name = input("Название системы: ")
print("Если хотите полного рандома в любой при очередном вводе напишите '-2'")
print("Если хотите локального рандома то напишите '-1'")
n_star = CountStar()
wb_template = load_workbook('template.xlsx')  # загрузка шаблона
wb = wb_template  # копирование шаблона
wsys = wb['Система']  # делаем активным основной лист
for i in range(1, n_star+1):

    if i > 25:    # литера для работы с файлом
        wlitter = chr(i // 26 + 64) + chr(i - (i // 26) * 26 + 65)
    else:
        wlitter = chr(i + 65)
    if i-1 > 25:  # литера для названия
        litter = chr((i-1) // 26 + 64) + chr((i-1) - ((i-1) // 26) * 26 + 65)
    else:
        litter = chr((i-1) + 65)
    if n_star == 1:
        name_star = name+' '
    else:
        name_star = name+' '+litter

    type_star = TypeStar(name_star)
    klass_star = KlassStar(name_star, type_star)
    n_planet = CountPlanet(name_star)
    color_star = ColorStar(klass_star)
    temp_star = TempStar(klass_star)
    rad_star_sun, rad_star_m = RadiusStar(klass_star)
    mass_star_sun, mass_star_kg = MassStar(klass_star)
    luminosity_vt = 4 * math.pi * rad_star_m ** 2 * PSB * temp_star ** 4
    luminosity_sun = luminosity_vt / LuminositySun
    wsys[wlitter + '1'] = name_star
    wsys[wlitter + '2'] = color_star
    wsys[wlitter + '3'] = TypesStar[type_star-1]
    wsys[wlitter + '4'] = klass_star
    wsys[wlitter + '5'] = n_planet
    wsys[wlitter + '6'] = temp_star
    wsys[wlitter + '7'] = rad_star_m
    wsys[wlitter + '8'] = rad_star_sun
    wsys[wlitter + '9'] = mass_star_kg
    wsys[wlitter + '10'] = mass_star_sun
    wsys[wlitter + '11'] = luminosity_vt
    wsys[wlitter + '12'] = luminosity_sun
    if n_planet > 0:
        wstar = wb.copy_worksheet(wb_template['Звезда'])
        wstar.title = name_star
        for j in range(1, 13):
            wstar['B'+str(j)].value = wsys[wlitter + str(j)].value
        ven_zone1 = (luminosity_sun ** 0.5) * 0.387
        ven_zone2 = (luminosity_sun ** 0.5) * 0.855
        life_zone1 = (luminosity_sun ** 0.5) * 0.855
        life_zone2 = (luminosity_sun ** 0.5) * 1.5
    orbit_planet_ae = 0
    for j in range(1, n_planet+1):

        if j > 24:  # литера для работы с файлом
            wplitter = chr((j + 1) // 26 + 64) + chr(j + 1 - ((j + 1) // 26) * 26 + 65)
        else:
            wplitter = chr(j + 66)
        if j - 1 > 25:  # литера для названия
            plitter = chr((j - 1) // 26 + 64) + chr((j - 1) - ((j - 1) // 26) * 26 + 65)
        else:
            plitter = chr((j - 1) + 65)
        plitter = plitter.lower()

        name_planet = name_star + plitter
        orbit_planet_ae, orbit_planet_m = PlanetOrbit(orbit_planet_ae)
        klass_planet = KlassPlanet(ven_zone1, ven_zone2, life_zone1, life_zone2, orbit_planet_ae, name_planet)
        type_planet = TypePlanet(klass_planet, name_planet)
        vplanet = (G * mass_star_kg / orbit_planet_m) ** 0.5
        year_planet = 2 * math.pi * orbit_planet_m / vplanet / 3600 / 24
        rings, n_moon, n_big_moon = CountMoon(type_planet, name_planet)
        intensity_planet = luminosity_vt / 4 / math.pi / (orbit_planet_m ** 2)
        density_planet, mass_planet_earth, mass_planet_kg = MassPlanet(type_planet, mass_star_sun)
        rad_planet_m = (3 * mass_planet_kg / 4 / math.pi / density_planet) ** (1/3)
        rad_planet_earth = rad_planet_m / RadiusEarth
        aof_planet = G * mass_planet_kg / rad_planet_m /rad_planet_m
        v1_planet = (aof_planet * rad_planet_m) ** 0.5
        v2_planet = 2 ** 0.5 * v1_planet
        v3_planet = ((2 ** 0.5 - 1) ** 2 * vplanet ** 2 + v2_planet ** 2) ** 0.5
        ConstSystem = random.randint(150, 250) / 100
        wstar[wplitter + '14'] = name_planet
        wstar[wplitter + '15'] = type_planet
        wstar[wplitter + '16'] = klass_planet
        wstar[wplitter + '17'] = n_moon
        wstar[wplitter + '18'] = n_big_moon
        wstar[wplitter + '19'] = round(rad_planet_m)
        wstar[wplitter + '20'] = round(rad_planet_earth, 4)
        wstar[wplitter + '21'] = mass_planet_kg
        wstar[wplitter + '22'] = mass_planet_earth
        wstar[wplitter + '23'] = density_planet
        wstar[wplitter + '24'] = round(aof_planet, 2)
        wstar[wplitter + '25'] = round(intensity_planet)
        wstar[wplitter + '26'] = round(orbit_planet_m)
        wstar[wplitter + '27'] = round(orbit_planet_ae, 3)
        wstar[wplitter + '30'] = round(year_planet, 2)
        wstar[wplitter + '31'] = round(v1_planet)
        wstar[wplitter + '32'] = round(v2_planet)
        wstar[wplitter + '33'] = round(v3_planet)
        wstar[wplitter + '34'] = round(vplanet)
        wstar[wplitter + '35'] = random.randint(0, 1800) / 10
        wstar[wplitter + '36'] = rings
        for z in range(1, n_big_moon+1):
            for f in range(0, len(StringNames)):
                wstar['A' + str(f+(z-1)*18+38)] = StringNames[f]
                wstar.cell(row=f+(z-1)*18+38, column=1).fill = myFill
            name_moon = name_planet + ' ' + str(z)
            type_moon = TypeMoon(ven_zone1, ven_zone2, life_zone1, life_zone2, orbit_planet_ae)
            list_mass_moon_earth = [0] * 5
            list_mass_moon_earth[0] = random.randint(96, min(8040, mass_planet_earth/4 * 10 ** 6)) / 10 ** 6
            list_mass_moon_earth[1] = random.randint(8040, min(25126, mass_planet_earth / 4 * 10 ** 6)) / 10 ** 6
            list_mass_moon_earth[2], list_mass_moon_earth[3] = list_mass_moon_earth[1], list_mass_moon_earth[1]
            list_mass_moon_earth[4] = random.randint(25126, min(1900000, mass_planet_earth / 4 * 10 ** 6)) / 10 ** 6
            mass_moon_earth = random.choice(list_mass_moon_earth)
            mass_moon_kg = mass_moon_earth * MassEarth
            density_moon = random.randint(110, 700) * 10
            rad_moon_m = (3 * mass_moon_kg / 4 / math.pi / density_moon) ** (1 / 3)
            rad_moon_earth = rad_moon_m / RadiusEarth
            aof_moon = G * mass_moon_kg / rad_moon_m / rad_moon_m
            v1_moon = (aof_moon * rad_moon_m) ** 0.5
            v2_moon = 2 ** 0.5 * v1_moon
            if aof_moon < 1 and (type_moon == 'Водный' or type_moon == 'Землеподобный' or type_moon == 'Амиачный'):
                type_moon = 'Насыщенный металлами'
            if mass_moon_earth < mass_planet_earth * 0.1167:
                lvloi = 'Приливной захват'
            else:
                lvloi = 'Приливная блокировка'
            pivo = (z-1)*18+38
            wstar[wplitter + str(0 + pivo)] = name_moon
            wstar[wplitter + str(1 + pivo)] = type_moon
            wstar[wplitter + str(2 + pivo)] = round(rad_moon_m)
            wstar[wplitter + str(3 + pivo)] = rad_moon_earth
            wstar[wplitter + str(4 + pivo)] = mass_moon_kg
            wstar[wplitter + str(5 + pivo)] = mass_moon_earth
            wstar[wplitter + str(6 + pivo)] = density_moon
            wstar[wplitter + str(7 + pivo)] = aof_moon
            wstar[wplitter + str(8 + pivo)] = '=('+ str(G*mass_planet_kg/4/(math.pi ** 2)) + '*(' \
                                                           + wplitter + str(11 + pivo) + '*86400)^2'\
                                                           + ')^(1/3)'
            wstar[wplitter + str(9 + pivo)] = '=' + wplitter + str(8 + pivo) + '/' + str(AE)
            wstar[wplitter + str(10 + pivo)] = '=' + wplitter + '29'
            if lvloi == 'Приливная блокировка':
                wstar[wplitter + str(11 + pivo)] = '=' + wplitter + '29'
            else:
                wstar[wplitter + str(11 + pivo)] = '=' + wplitter + '29' + '*' + str(ConstSystem ** z)
            wstar[wplitter + str(12 + pivo)] = lvloi
            wstar[wplitter + str(13 + pivo)] = v1_moon
            wstar[wplitter + str(14 + pivo)] = v2_moon
            wstar[wplitter + str(15 + pivo)] = '=2*' + str(math.pi) + '*' \
                                                            + wplitter + str(8 + pivo) +'/86400/' \
                                                            + wplitter + str(11 + pivo)
            wstar[wplitter + str(16 + pivo)] = random.randint(0, 1800) / 10


del wb["Звезда"]
wb.worksheets[0].name = "Main"
wb.save(name+'.xlsx')
